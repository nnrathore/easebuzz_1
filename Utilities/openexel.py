import openpyxl
from openpyxl import Workbook

wb = openpyxl.load_workbook("first_test.xlsx")
ws = wb.active

wss = wb.sheetnames
print(wss[1])
sheet1 = wb["Sheet1"]
print(sheet1.title)
ws.cell(2,5,"test")
print(ws.cell(2,5).value)
print(ws.max_row)
print(ws.max_column)
print(ws.min_row)
print(ws.min_column)
ws.merge_cells('A1:A5')
wb.save("first_test.xlsx")
wb.close()